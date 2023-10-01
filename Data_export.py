import os
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def Excel_export(df, Export_Path, Report_Filename, id):

    # Load existing sheet as it is
    book = load_workbook(os.path.join(Export_Path, Report_Filename))

    # create a new sheet
    Sheet_name = str(id)
    sheet = book.create_sheet(Sheet_name)

    # Load dataframe into new sheet
    for row in dataframe_to_rows(df, index=False, header=True):
        sheet.append(row)

    # Save the modified excel at desired location
    book.save(os.path.join(Export_Path, Report_Filename))
